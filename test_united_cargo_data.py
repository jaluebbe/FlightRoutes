import datetime
import logging
import pathlib
import tempfile
import arrow
import openpyxl
import pytest
from unittest.mock import patch
from united_cargo_data import (
    _operating_days,
    _parse_int_time,
    _parse_time_obj,
    _process_widebody_row,
    _process_simple_row,
    Airline,
)

# ---------------------------------------------------------------------------
# XLSX fixture helpers
# ---------------------------------------------------------------------------

# Monday 2026-04-06 — isoweekday() == 1
_MON = arrow.get("2026-04-06T12:00:00")
# Thursday 2026-04-09
_THU = arrow.get("2026-04-09T12:00:00")

_APR_START = datetime.datetime(2026, 4, 1)
_APR_END = datetime.datetime(2026, 4, 30)


def _make_xlsx(
    widebody_rows: list[list],
    narrowbody_rows: list[list],
    uax_rows: list[list],
) -> pathlib.Path:
    """Build a minimal United Cargo XLSX with three sheets and return its path
    in a temp directory."""
    wb = openpyxl.Workbook()

    # Widebody sheet — no fixed header, section headers interspersed.
    ws_wb = wb.active
    ws_wb.title = "Widebody"
    # Preamble rows (rows 0-7 in 0-based, i.e. Excel rows 1-8)
    for _ in range(8):
        ws_wb.append([""] * 11)
    # Column header row (Excel row 9, 0-based index 8)
    ws_wb.append(
        [
            "Sales region",
            "Org",
            "Des",
            "Flight #",
            "Eff Date",
            "Dis Date",
            "Departs",
            "Arrives",
            "A/C type",
            "DOW",
            "Notes",
        ]
    )
    for _row in widebody_rows:
        ws_wb.append(_row)

    # Narrowbody sheet — fixed header at row index 7 (Excel row 8).
    ws_nb = wb.create_sheet("Narrowbody")
    for _ in range(7):
        ws_nb.append([""] * 7)
    ws_nb.append(
        ["Org", "Des", "Flight #", "Departs", "Arrives", "A/C type", "DOW"]
    )
    for _row in narrowbody_rows:
        ws_nb.append(_row)

    # UAX sheet — same structure as Narrowbody.
    ws_uax = wb.create_sheet("UAX")
    for _ in range(7):
        ws_uax.append([""] * 7)
    ws_uax.append(
        ["Org", "Des", "Flight #", "Departs", "Arrives", "A/C type", "DOW"]
    )
    for _row in uax_rows:
        ws_uax.append(_row)

    # JV sheet — present but unused.
    wb.create_sheet("JV")

    _tmp = tempfile.mkdtemp()
    _path = (
        pathlib.Path(_tmp)
        / "United_Cargo_Flight_Schedule-Full_Network-April2026.xlsx"
    )
    wb.save(_path)
    return _path


def _run_update(xlsx_path: pathlib.Path, utc: arrow.Arrow) -> list[dict]:
    stored = []
    airline = Airline()
    with patch("united_cargo_data.PWD", xlsx_path.parent):
        with patch.object(airline, "update_flight", side_effect=stored.append):
            airline.update_data(utc)
    return stored


# ---------------------------------------------------------------------------
# _operating_days
# ---------------------------------------------------------------------------


def test_operating_days_all():
    assert _operating_days("1234567") == set(range(1, 8))


def test_operating_days_single():
    assert _operating_days(1) == {1}


def test_operating_days_subset():
    assert _operating_days(135) == {1, 3, 5}


# ---------------------------------------------------------------------------
# _parse_int_time
# ---------------------------------------------------------------------------


def test_parse_int_time_basic():
    result = _parse_int_time(1430, "2026-04-06", "America/New_York")
    assert result.hour == 14
    assert result.minute == 30


def test_parse_int_time_zero_padded():
    result = _parse_int_time(545, "2026-04-06", "Europe/London")
    assert result.hour == 5
    assert result.minute == 45


# ---------------------------------------------------------------------------
# _parse_time_obj
# ---------------------------------------------------------------------------


def test_parse_time_obj_basic():
    _t = datetime.time(9, 30)
    result = _parse_time_obj(_t, "2026-04-06", "America/Chicago")
    assert result.hour == 9
    assert result.minute == 30


# ---------------------------------------------------------------------------
# Widebody rows
# ---------------------------------------------------------------------------

_WB_ROW_DAILY = [
    "U.S. East",
    "JFK",
    "LHR",
    "UA 0008",
    _APR_START,
    _APR_END,
    1930,
    755,
    "767",
    1234567,
    None,
]

_WB_ROW_MON_ONLY = [
    "U.S. East",
    "EWR",
    "FRA",
    "UA 0018",
    _APR_START,
    _APR_END,
    1800,
    815,
    "767",
    1,
    None,
]

_WB_ROW_SHORT_VALIDITY = [
    "U.S. East",
    "LAX",
    "NRT",
    "UA 0837",
    datetime.datetime(2026, 4, 6),
    datetime.datetime(2026, 4, 6),
    1100,
    1430,
    "789",
    1234567,
    None,
]


def test_widebody_daily_flight_stored():
    _path = _make_xlsx([_WB_ROW_DAILY], [], [])
    stored = _run_update(_path, _MON)
    assert len(stored) == 1
    _f = stored[0]
    assert _f["airline_iata"] == "UA"
    assert _f["airline_icao"] == "UAL"
    assert _f["flight_number"] == 8
    assert _f["route"] == "KJFK-EGLL"


def test_widebody_wrong_day_skipped():
    _path = _make_xlsx([_WB_ROW_MON_ONLY], [], [])
    stored = _run_update(_path, _THU)
    assert stored == []


def test_widebody_correct_day_stored():
    _path = _make_xlsx([_WB_ROW_MON_ONLY], [], [])
    stored = _run_update(_path, _MON)
    assert len(stored) == 1


def test_widebody_outside_validity_skipped():
    _path = _make_xlsx([_WB_ROW_SHORT_VALIDITY], [], [])
    stored = _run_update(_path, _THU)
    assert stored == []


def test_widebody_within_validity_stored():
    _path = _make_xlsx([_WB_ROW_SHORT_VALIDITY], [], [])
    stored = _run_update(_path, _MON)
    assert len(stored) == 1


def test_widebody_overnight_arrival_shifted():
    # Dep 23:00, arr 07:00 — should shift arrival to next day
    _row = [
        "U.S. East",
        "JFK",
        "LHR",
        "UA 0008",
        _APR_START,
        _APR_END,
        2300,
        700,
        "767",
        1234567,
        None,
    ]
    _path = _make_xlsx([_row], [], [])
    stored = _run_update(_path, _MON)
    assert len(stored) == 1
    assert stored[0]["arrival"] > stored[0]["departure"]


def test_widebody_flight_id_format():
    _path = _make_xlsx([_WB_ROW_DAILY], [], [])
    stored = _run_update(_path, _MON)
    assert stored[0]["_id"] == "UA_8_KJFK-EGLL_20260406"


# ---------------------------------------------------------------------------
# Narrowbody / UAX rows (datetime.time objects)
# ---------------------------------------------------------------------------

_NB_ROW_DAILY = [
    "CLT",
    "ATL",
    1557,
    datetime.time(12, 49),
    datetime.time(14, 5),
    "319",
    1234567,
]

_NB_ROW_THU_ONLY = [
    "ORD",
    "DFW",
    200,
    datetime.time(8, 0),
    datetime.time(10, 30),
    "738",
    4,
]


def test_narrowbody_daily_stored():
    _path = _make_xlsx([], [_NB_ROW_DAILY], [])
    stored = _run_update(_path, _MON)
    assert len(stored) == 1
    assert stored[0]["flight_number"] == 1557


def test_narrowbody_wrong_day_skipped():
    _path = _make_xlsx([], [_NB_ROW_THU_ONLY], [])
    stored = _run_update(_path, _MON)
    assert stored == []


def test_narrowbody_correct_day_stored():
    _path = _make_xlsx([], [_NB_ROW_THU_ONLY], [])
    stored = _run_update(_path, _THU)
    assert len(stored) == 1


def test_uax_row_stored():
    _uax_row = [
        "ABE",
        "ORD",
        5913,
        datetime.time(17, 45),
        datetime.time(19, 5),
        "CRJ",
        1234567,
    ]
    _path = _make_xlsx([], [], [_uax_row])
    stored = _run_update(_path, _MON)
    assert len(stored) == 1
    assert stored[0]["airline_icao"] == "UAL"


def test_both_widebody_and_narrowbody_combined():
    _path = _make_xlsx([_WB_ROW_DAILY], [_NB_ROW_DAILY], [])
    stored = _run_update(_path, _MON)
    assert len(stored) == 2


def test_missing_file_logs_warning(caplog, tmp_path):
    airline = Airline()
    stored = []
    with caplog.at_level(logging.WARNING, logger="united_cargo_data.py"):
        with patch("united_cargo_data.PWD", tmp_path):
            with patch.object(
                airline, "update_flight", side_effect=stored.append
            ):
                airline.update_data(_MON)
    assert stored == []
    assert any("not found" in _r.message for _r in caplog.records)


def test_smoke_active_flights():
    airline = Airline()
    utc = arrow.utcnow().timestamp()
    flights = airline.get_active_flights(utc)
    assert isinstance(flights, list)
